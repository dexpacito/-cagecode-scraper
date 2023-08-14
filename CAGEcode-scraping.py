'''
works with template from cap list file, check column order and indexing
when appending data. requires to install libraries.

working as of 07/23
'''

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import openpyxl

driver = webdriver.Chrome()

webpage_url = "https://cage.dla.mil/Search"
driver.get(webpage_url)

# Wait for the "I Agree" button to be clickable and click
agree_button = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.XPATH, "//input[@type='submit' and @value='I Agree']"))
)

agree_button.click()

# Load the Excel file
excel_file_path = '/Users/dextersrstlne/Desktop/siaec/forscraping.xlsx'  #replace this with actual file path
workbook = openpyxl.load_workbook(excel_file_path)
sheet = workbook.active

# Iterate through the column in Excel
for row in range(1, sheet.max_row + 1):
    # Get the value from the current cell in the column
    cell_value = sheet.cell(row=row, column=1).value

    # Find search bar on search page
    search_bar = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.ID, "SearchString"))
    )

    # Clear search bar and enter code into it
    search_bar.clear()
    search_bar.send_keys(cell_value)
    search_bar.send_keys(Keys.RETURN)

    # Enter Details page
    details_link = WebDriverWait(driver, 15).until(
        EC.element_to_be_clickable((By.XPATH, "//tr//td/a[contains(text(),'Details')]"))
    )

    details_link.click()

    try:
        address_div = WebDriverWait(driver,15).until(
            EC.visibility_of_element_located((By.XPATH,"//div[label[text()='Address']]"))
        )

        address_span = address_div.find_element(By.TAG_NAME,"span")
        address_text = address_span.text

        #Write address to column D in the same row
        sheet.cell(row=row, column=4).value = address_text

    except:
        print(f'Address not found for row {row}')
        continue

    try:
        postal_div = WebDriverWait(driver, 15).until(
            EC.visibility_of_element_located((By.XPATH, "//div[label[text()='Zip/Postal']]"))
        )

        postal_span = postal_div.find_element(By.TAG_NAME, "span")
        postal_text = postal_span.text

        # Write the postal to column E in the same row
        sheet.cell(row=row, column=5).value = postal_text

    except:
        print(f"Postal not found for row {row}")
        continue

    try:
        phone_div = WebDriverWait(driver, 15).until(
            EC.visibility_of_element_located((By.XPATH,"//div[label[text()='Phone']]"))
        )

        phone_span = phone_div.find_element(By.TAG_NAME, "span")
        phone_text = phone_span.text

        # Write the phone number to column H in the same row
        sheet.cell(row=row, column = 8).value = phone_text

    except:
        print(f"Phone number not found for row {row}")

    driver.get(webpage_url)

workbook.save(excel_file_path)

driver.quit()